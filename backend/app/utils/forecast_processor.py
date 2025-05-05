from datetime import datetime
import pandas as pd
import openpyxl
import os
import shutil
import numpy as np
from openpyxl.utils import get_column_letter, column_index_from_string
pd.set_option('future.no_silent_downcasting', True) 

def procesar_forecast( epm_file_path, forecast_base_path):
    print("üìÅ Iniciando limpieza del archivo EPM...")
    # Paso 1: Leer archivo EPM
    try:
        print(f"üìÑ Leyendo archivo EPM desde: {epm_file_path}")
        epm_df_raw = pd.read_excel(epm_file_path, engine='openpyxl', header=None)
        print("‚úÖ Archivo EPM le√≠do con √©xito.")
    except Exception as e:
        print(f"‚ö†Ô∏è Error al leer el archivo EPM: {e}")
        return None
    
    # Paso 2: Eliminar primeras 3 filas
    epm_df = epm_df_raw.iloc[3:].reset_index(drop=True)
    print("üßπ Primeras 3 filas eliminadas.")
    
    # Paso 3: Eliminar desde la fila "Overall - Total"
    print("üîç Buscando fila 'Overall - Total'...")
    for idx, row in enumerate(epm_df.iloc[:, :10].values):
        if any('Overall - Total' in str(cell) for cell in row):
            epm_df = pd.concat([epm_df.iloc[:idx], epm_df.iloc[idx+3:]]).reset_index(drop=True)
            print(f"üóëÔ∏è Filas desde 'Overall - Total' eliminadas (√≠ndice {idx}).")
            break
    else:
        print("‚ö†Ô∏è No se encontr√≥ la fila 'Overall - Total'. Se contin√∫a sin eliminar.")
    
    # Paso 4: Eliminar columnas espec√≠ficas por √≠ndice
    columnas_a_eliminar = [9, 80, 86, 87]
    epm_df.drop(epm_df.columns[columnas_a_eliminar], axis=1, inplace=True)
    print("üóëÔ∏è Columnas J, CB, CH y CI eliminadas. Total de columnas ahora: {}".format(epm_df.shape[1]))
    
    # Paso 5: Asignar encabezados
    epm_df.columns = epm_df.iloc[0]
    epm_df = epm_df[1:].reset_index(drop=True)
    print("‚úÖ Encabezados reasignados desde el propio archivo EPM.")
    
    # Paso 6: Preparar directorios
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    temp_dir = os.path.join(base_dir, 'temp')
    os.makedirs(temp_dir, exist_ok=True)
    
    # Crear archivo limpio
    output_path = os.path.join(temp_dir, 'epm_limpio.xlsx')
    epm_df.to_excel(output_path, index=False)
    print(f"üíæ Archivo limpio guardado en: {output_path}")
    
    # Paso 7: Crear copia temporal del archivo COPY base
    # Asegurarse que forecast_base_path tiene la extensi√≥n correcta (.xlsm)
    forecast_base_dir = os.path.dirname(forecast_base_path)
    forecast_base_name = os.path.splitext(os.path.basename(forecast_base_path))[0]
    
    # Corregir la extensi√≥n si es necesario
    forecast_base_path_corrected = os.path.join(forecast_base_dir, f"{forecast_base_name}.xlsm")
    if not os.path.exists(forecast_base_path_corrected):
        print(f"‚ö†Ô∏è Archivo no encontrado: {forecast_base_path_corrected}")
        # Intentar con extensi√≥n .xlsx como fallback
        alternative_path = os.path.join(forecast_base_dir, f"{forecast_base_name}.xlsx")
        if os.path.exists(alternative_path):
            forecast_base_path_corrected = alternative_path
            print(f"‚úÖ Usando archivo alternativo: {forecast_base_path_corrected}")
        else:
            print("‚ùå No se encontr√≥ el archivo forecast_base con ninguna extensi√≥n.")
            return None
    
    # Mantener copy_base.xlsx como est√°
    copy_path = os.path.join(os.path.dirname(forecast_base_path_corrected), 'copy_base.xlsx')
    temp_copy_path = os.path.join(temp_dir, 'copy_base_temp.xlsx')
    
    if not os.path.exists(copy_path):
        print(f"‚ö†Ô∏è Archivo copy_base no encontrado: {copy_path}")
        # Intentar con extensi√≥n .xlsm como fallback
        copy_path_alt = os.path.join(os.path.dirname(forecast_base_path_corrected), 'copy_base.xlsm')
        if os.path.exists(copy_path_alt):
            copy_path = copy_path_alt
            print(f"‚úÖ Usando archivo copy_base alternativo: {copy_path}")
        else:
            print("‚ùå No se encontr√≥ el archivo copy_base con ninguna extensi√≥n.")
            return None
    
    shutil.copy(copy_path, temp_copy_path)
    print(f"üìÇ Copia temporal creada: {temp_copy_path}")
    
    # Cargar el workbook COPY
    print("üìä Cargando archivo COPY para procesamiento optimizado...")
    wb = openpyxl.load_workbook(temp_copy_path)
    ws_row = wb['Row']
    ws_pivot = wb['Pivot']
    
    # Mapeo columnas Pivot <- Row
    col_mapping = {
        'B': 'R', 'C': 'B', 'D': 'E', 'E': 'AL', 'F': 'AZ', 'G': 'AF',
        'H': 'J', 'I': 'BJ', 'J': 'K', 'K': 'L', 'L': 'O', 'M': 'AJ',
        'N': 'S', 'O': 'BB', 'P': 'BA', 'Q': 'AY', 'R': 'T', 'S': 'BW',
        'T': 'CE', 'U': 'Q', 'V': 'CG', 'W': 'D', 'X': 'BX', 'Y': 'BL',
        'Z': 'AE', 'AA': 'V', 'AB': 'X', 'AC': 'W', 'AD': 'U'
    }
    
    print("üìä Preparando datos para optimizaci√≥n...")
    # 1. Datos para hoja Row
    data_values = epm_df.values
    max_cols = 200
    row_data = np.empty((len(data_values), max_cols), dtype=object)
    for c_idx, col in enumerate(epm_df.columns):
        col_index = c_idx + 1
        if col_index < max_cols:
            col_values = epm_df[col].fillna('').infer_objects(copy=False).values # evita NaN que generen desfase
            row_data[:, col_index] = col_values

    
    batch_size = 1000
    total_rows = len(data_values)
    for start_idx in range(0, total_rows, batch_size):
        end_idx = min(start_idx + batch_size, total_rows)
        batch = row_data[start_idx:end_idx]
        for r_idx, row in enumerate(batch, start=start_idx+2):
            for c_idx in range(1, max_cols):
                val = row[c_idx]
                if pd.notnull(val):
                    ws_row.cell(row=r_idx, column=c_idx+1, value=val)
    
    print("‚úÖ Datos optimizados en hoja 'Row'.")
    
    # 2. Datos para hoja Pivot
    print("üîÑ Optimizando transferencia a hoja 'Pivot'...")
    pivot_start_row = 4
    col_indices = {dest: column_index_from_string(src) for dest, src in col_mapping.items()}
    pivot_data = np.empty((total_rows, len(col_mapping)), dtype=object)
    
    for r_idx in range(total_rows):
        for col_idx, (dest_col, src_col_idx) in enumerate(col_indices.items()):
            if src_col_idx - 1 < max_cols:
                val = row_data[r_idx, src_col_idx - 1]
                if pd.notnull(val):
                    pivot_data[r_idx, col_idx] = val
    
    # Filtrar y ordenar datos antes de pasarlos a Pivot
    print("üîç Aplicando filtros a datos de Pivot...")
    pivot_df = pd.DataFrame(pivot_data, columns=list(col_mapping.keys()))
    
    # Filtros seg√∫n encabezados en Pivot
    allowed_countries = [
        "Colombia", "Costa Rica", "Dominican Republic", "El Salvador",
        "Guatemala", "Honduras", "Nicaragua", "Panama", "Venezuela"
    ]
    allowed_ut_lvl_15 = [
        "Mainframe SW", "Power", "Storage", "Z HW", "Storage HW",
        "Power HW", "Storage TPS", "Z TPS", "Power TPS"
    ]
    
    # Aplica filtros (col H ‚Üí 'H', col K ‚Üí 'K', col B ‚Üí 'B')
    pivot_df_filtered = pivot_df[
        pivot_df['H'].isin(allowed_countries) &
        pivot_df['K'].isin(allowed_ut_lvl_15)
    ]
    
    # Ordenar por Customer Name (columna B)
    pivot_df_filtered = pivot_df_filtered.sort_values(by='B')
    print(f"‚úÖ Datos filtrados: {len(pivot_df_filtered)} filas conservadas.")
    
    # Copiar al Excel filtrado
    for r_idx, row in enumerate(pivot_df_filtered.values):
        for col_idx, dest_col in enumerate(col_mapping.keys()):
            dest_col_idx = column_index_from_string(dest_col)
            val = row[col_idx]
            if pd.notnull(val):
                ws_pivot.cell(row=r_idx + pivot_start_row, column=dest_col_idx, value=val)

    # Guardar copy_base_temp.xlsx
    wb.save(temp_copy_path)
    print("‚úÖ COPY base actualizado.")
    
    # 3. Crear forecast temporal y copiar datos Pivot ‚Üí EPM
    print("üì§ Copiando datos de Pivot a forecast temporal...")
    
    # Crear copia del forecast con extensi√≥n xlsm
    forecast_temp_path = os.path.join(temp_dir, 'Systems HW - North SSA EPM ISC.xlsm')
    shutil.copy(forecast_base_path_corrected, forecast_temp_path)
    
    # Cargar con openpyxl manteniendo VBA
    forecast_wb = openpyxl.load_workbook(forecast_temp_path, keep_vba=True)
    forecast_epm_ws = forecast_wb['EPM']
    
    # Determinar la √∫ltima fila con datos en el rango B:AD de Pivot
    pivot_range_start = column_index_from_string('B')
    pivot_range_end = column_index_from_string('AD')
    last_data_row = pivot_start_row
    
    for col in range(pivot_range_start, pivot_range_end + 1):
        for row in range(pivot_start_row, ws_pivot.max_row + 1):
            if ws_pivot.cell(row=row, column=col).value is not None:
                last_data_row = max(last_data_row, row)
    
    print(f"üî¢ √öltima fila con datos en Pivot (rango B:AD): {last_data_row}")
    
    # Copiar de Pivot!B4:AD ‚Üí Forecast!B4:AD para todas las filas con datos
    pivot_cols = range(pivot_range_start, pivot_range_end + 1)
    forecast_cols = range(column_index_from_string('B'), column_index_from_string('AD') + 1)
    
    # Copiar todas las filas desde pivot_start_row hasta last_data_row
    for r in range(pivot_start_row, last_data_row + 1):
        for src_col, dest_col in zip(pivot_cols, forecast_cols):
            val = ws_pivot.cell(row=r, column=src_col).value
            # Copiamos incluso valores nulos para mantener la estructura
            forecast_epm_ws.cell(row=r, column=dest_col, value=val)
        # 3. Copiar columnas seleccionadas de EPM a hoja Data Trx 

    print("üìå Copiando columnas clave desde hoja 'EPM' a hoja 'Data Trx'...")

    try:
        ws_epm_forecast = forecast_wb['EPM']
        ws_DataTrx = forecast_wb['Data Trx']

    # Primero, llenamos la columna AI (Brand) manualmente en EPM
        def calcular_brand(valor):
            if valor == "Storage HW":
             return "Storage HW"
            elif valor == "Storage TPS":
                return "Storage TPS"
            elif valor == "Z HW":
             return "Mainframe"
            elif valor == "Z TPS":
                return "Z Middleware"
            elif valor == "Power HW":
                return "Cognitive"
            elif valor == "Power TPS":
                return "Cognitive"
            return None

        print("üîç Calculando valores de la columna AI (Brand)...")
        for row in range(4, ws_epm_forecast.max_row + 1):
            valor_lvl15 = ws_epm_forecast.cell(row=row, column=11).value  # Columna AH
            brand = calcular_brand(valor_lvl15)
            ws_epm_forecast.cell(row=row, column=35).value = brand  # Columna AI
        print("‚úÖ Columna AI (Brand) calculada y completada.")

        # Definimos la tabla de correspondencia para pa√≠ses
        tabla_country = {
            "Colombia": "Colombia",
            "Costa Rica": "LCR",
            "Dominican Republic": "LCR",
            "El Salvador": "LCR",
            "Guatemala": "LCR",
            "Nicaragua": "LCR",
            "Honduras": "LCR",
            "Panama": "LCR",
            "Venezuela": "Venezuela"
        }

        def calcular_country(nombre_pais):
            if nombre_pais is None:
                return ""
            nombre_pais_str = str(nombre_pais).strip()
            return tabla_country.get(nombre_pais_str, "")  # Devuelve "" si no encuentra el pa√≠s

        # üîç Calcular columna AJ (Country) en EPM
        print("üåé Calculando valores de la columna AJ (Country) en EPM...")
        for row in range(4, ws_epm_forecast.max_row + 1):
            country_name = ws_epm_forecast.cell(row=row, column=8).value  # Columna H (Country Name)
            resultado = calcular_country(country_name)
            ws_epm_forecast.cell(row=row, column=36).value = resultado  # Columna AJ (Country)
        print("‚úÖ Columna AJ (Country) en EPM calculada.")

        def extraer_mes(forecast_month):
            """
            Extrae el car√°cter en la posici√≥n 7 (similar a =EXTRAE(texto;7;1) en Excel)

            Args:
                forecast_month: El valor de la celda Forecast Month

            Returns:
                El car√°cter en la posici√≥n 7 (√≠ndice 6 en Python) o vac√≠o si no existe
            """
            if forecast_month is None or not isinstance(forecast_month, str):
                return ""

            # En Python, los √≠ndices empiezan desde 0, as√≠ que la posici√≥n 7 es el √≠ndice 6
            if len(forecast_month) >= 7:
                return forecast_month[6:7]  # Extrae 1 car√°cter desde la posici√≥n 7
            else:
                return ""

        # Ejemplo de uso en tu c√≥digo:
        print("üîç Extrayendo valores de mes...")
        for row in range(4, ws_epm_forecast.max_row + 1):
            forecast_month = ws_epm_forecast.cell(row=row, column=6).value  
            mes = extraer_mes(forecast_month)
            ws_epm_forecast.cell(row=row, column=44).value = mes

        def calcular_etiqueta(fila):
            try:
                if fila ['Sum of Call'] != 0:
                    return "Call"
                elif fila['Sum of Upside'] != 0:
                    return "Upside"
                else:
                    return "Stretch"
            except Exception:
                return ""

        print("üîç Calculando etiqueta (Call, Upside, Stretch)...")
        for row in range(4, ws_epm_forecast.max_row + 1):
            sum_call = ws_epm_forecast.cell(row=row, column=27).value  # columna 'Sum of Call'
            sum_upside = ws_epm_forecast.cell(row=row, column=29).value  # columna 'Sum of Upside'

             #Creamos un "diccionario" con los datos esperados por la funci√≥n
            fila = {
                'Sum of Call': sum_call or 0,       # En caso de que venga None, usamos 0
                'Sum of Upside': sum_upside or 0
            }

            etiqueta = calcular_etiqueta(fila)
            ws_epm_forecast.cell(row=row, column=41).value = etiqueta  # columna donde escribimos la etiqueta



        def calcular_estado(rdmp, isc_sales_stage_name):
            """
            Implementa la f√≥rmula:
            =SI.ERROR(SI(Y(Data_ISC[@RDMP]="Call";Data_ISC[@[ISC Sales Stage Name]]<>"Won");"At Risk";
                        SI(Data_ISC[@[ISC Sales Stage Name]]="Won";"Won";
                        SI(Data_ISC[@RDMP]="Upside";"Key Stretch";"Stretch")));"")
    
            Args:
                rdmp: Valor de la columna RDMP
                isc_sales_stage_name: Valor de la columna ISC Sales Stage Name
    
            Returns:
                El resultado calculado seg√∫n la f√≥rmula
         """
            try:
            # Primera condici√≥n: Si RDMP es "Call" y ISC Sales Stage Name no es "Won"
                if rdmp == "Call" and isc_sales_stage_name != "Won":
                    return "At Risk"
                # Segunda condici√≥n: Si ISC Sales Stage Name es "Won"
                elif isc_sales_stage_name == "Won":
                    return "Won"
                # Tercera condici√≥n: Si RDMP es "Upside"
                elif rdmp == "Upside":
                    return "Key Stretch"
                # Caso predeterminado
                else:
                    return "Stretch"
            except Exception:
                # Manejo de cualquier error (equivalente a SI.ERROR)
                return ""
    

        # Ejemplo de uso en tu c√≥digo:
        print("üîç Calculando estado seg√∫n RDMP y ISC Sales Stage...")
        for row in range(4, ws_epm_forecast.max_row + 1):
            rdmp = ws_epm_forecast.cell(row=row, column=41).value  
            isc_stage = ws_epm_forecast.cell(row=row, column=7).value
    
            estado = calcular_estado(rdmp, isc_stage)
            ws_epm_forecast.cell(row=row, column=43).value = estado  


        def copiar_sum_of_oppty_value(ws, fila_origen, columna_origen, fila_destino, columna_destino):
            """
            Copia el valor de Sum of Oppty Value a otra celda
    
            Args:
                ws: Hoja de trabajo (worksheet)
                fila_origen: N√∫mero de fila donde leer el valor
                columna_origen: N√∫mero de columna donde est√° "Sum of Oppty Value"
                fila_destino: N√∫mero de fila donde escribir el valor
                columna_destino: N√∫mero de columna donde escribir el valor
            """
            valor = ws.cell(row=fila_origen, column=columna_origen).value
            ws.cell(row=fila_destino, column=columna_destino).value = valor

        # Ejemplo de uso:
        print("üîç Copiando valores de Sum of Oppty Value...")
        columna_sum_of_oppty = 30  # Reemplaza X con el n√∫mero de columna donde est√° "Sum of Oppty Value"
        columna_destino = 33  # Reemplaza Y con el n√∫mero de columna donde quieres guardar el resultado

        for row in range(4, ws_epm_forecast.max_row + 1):
            # Copiar el valor de la misma fila
            copiar_sum_of_oppty_value(
                ws_epm_forecast, 
                fila_origen=row, 
                columna_origen=columna_sum_of_oppty,
                fila_destino=row, 
                columna_destino=columna_destino
            )
        print("‚úÖ Valores de Sum of Oppty Value copiados.")

    # üì• Luego copiamos las columnas a Data Trx
        columnas_mapeo = {
            'AI': 'A',
            'AJ': 'B',
            'C': 'C',
            'N': 'D',
            'B': 'E',
            'D': 'F',
            'AR': 'G',
            'E': 'H',
            'S': 'I',
            'AQ': 'J',
            'AG': 'K'
        }

        start_row_epm = 4  # desde fila 4 en EPM
        start_row_DataTrx = 4   # fila 3 en Data Trx
        max_row = ws_epm_forecast.max_row
        
        
        for epm_col, DataTrx_col in columnas_mapeo.items():
            col_idx_epm = column_index_from_string(epm_col)
            col_idx_DataTrx = column_index_from_string(DataTrx_col)

            print(f"  üîÑ Copiando columna {epm_col} ‚ûú {DataTrx_col}")
            for i in range(start_row_epm, max_row + 1):
                value = ws_epm_forecast.cell(row=i, column=col_idx_epm).value
                target_row = start_row_DataTrx + (i - start_row_epm)
                ws_DataTrx.cell(row=target_row, column=col_idx_DataTrx, value=value)

        print("‚úÖ Columnas copiadas correctamente de EPM a P2.")

    except Exception as e:
        print(f"‚ùå Error al copiar datos de EPM a P2: {e}")
# üíæ Guardar con extensi√≥n xlsm
    forecast_wb.save(forecast_temp_path)
    print(f"‚úÖ Forecast temporal actualizado con todas las filas de datos: {forecast_temp_path}")

    if os.path.exists(output_path):
        os.remove(output_path)
        print(f"üóëÔ∏è Eliminado: {output_path}")

    if os.path.exists(temp_copy_path):
        os.remove(temp_copy_path)
        print(f"üóëÔ∏è Eliminado: {temp_copy_path}")

        # Paso 8: Limpiar la carpeta temporal
    temp_dir = 'temp' 
    print("üßπ Limpiando la carpeta temporal...")
    try:    
        for filename in os.listdir(temp_dir):
            file_path = os.path.join(temp_dir, filename)
            if os.path.isfile(file_path):
                os.remove(file_path)
                print(f"üóëÔ∏è Eliminado archivo: {file_path}")
        print("‚úÖ Todos los archivos de la carpeta 'temp' han sido eliminados.")
    except Exception as e:
        print(f"‚ö†Ô∏è Error al eliminar archivos de la carpeta 'temp': {e}")
    return forecast_temp_path