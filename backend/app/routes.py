from flask import Blueprint, request, send_file, jsonify
import os
import time
import logging
from openpyxl import load_workbook
import pandas as pd
from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from openpyxl.utils.dataframe import dataframe_to_rows

from .utils.forecast_processor import procesar_forecast

# Configuración de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('epm_automation.log'),
        logging.StreamHandler()
    ]
)

main = Blueprint('main', __name__)

# Configuración de rutas
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DOWNLOAD_FOLDER = os.path.join(BASE_DIR, 'temp')
os.makedirs(DOWNLOAD_FOLDER, exist_ok=True)
logging.info(f"Ruta de DOWNLOAD_FOLDER: {DOWNLOAD_FOLDER}")

# Configuración de Selenium
EDGE_DRIVER_PATH = os.path.join(BASE_DIR, 'WebDriver', 'msedgedriver.exe')
COGNOS_URL = "https://w3.ibm.com/epm/app-prod/bi/?perspective=content&tab=teamContent&folder=i47BFB972345742C5936DA23E6AE48D8D"
COGNOS_USER = os.getenv("COGNOS_USER", "Alexix.Orostegui@ibm.com")
COGNOS_PASS = os.getenv("COGNOS_PASS", "Orostegui1016952406")

def configure_edge():
    edge_options = EdgeOptions()
    edge_options.use_chromium = True  # Esto es obligatorio
    edge_options.add_argument("--start-fullscreen")
    edge_options.add_argument("--disable-notifications")
    edge_options.add_experimental_option("prefs", {
        "download.default_directory": DOWNLOAD_FOLDER,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True
    })
    return edge_options

from openpyxl.utils.dataframe import dataframe_to_rows
from flask import request, jsonify
import pandas as pd
from openpyxl import load_workbook
import os
import logging

@main.route('/process-budget', methods=['POST'])
def process_budget():
    try:
        print("1. Iniciando procesamiento de archivo budget...")
        
        # 1. Verificar que se envió el archivo
        print("2. Verificando si se recibió el archivo...")
        if 'budgetFile' not in request.files:
            print("Error: No se encontró 'budgetFile' en los archivos recibidos")
            return jsonify({"error": "No se proporcionó archivo"}), 400
        
        budget_file = request.files['budgetFile']
        print(f"3. Archivo recibido: {budget_file.filename}")
        
        # 2. Validar extensión del archivo
        print("4. Validando extensión del archivo...")
        if not budget_file.filename.lower().endswith(('.xlsx', '.xls', '.xlsm')):
            print(f"Error: Formato de archivo no válido: {budget_file.filename}")
            return jsonify({"error": "Formato de archivo no válido"}), 400
        
        # 3. Definir rutas
        print("5. Definiendo rutas...")
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        forecast_path = os.path.join(BASE_DIR, 'data', 'forecast_base.xlsm')
        print(f"6. Ruta base: {BASE_DIR}")
        print(f"7. Ruta forecast: {forecast_path}")
        
        # Verificar si existe el archivo forecast_base.xlsm
        if not os.path.exists(forecast_path):
            print(f"Error: No se encontró el archivo forecast_base en: {forecast_path}")
            return jsonify({"error": "Archivo forecast_base.xlsm no encontrado"}), 400
        
        # 4. Leer datos del archivo subido (hoja budget)
        print("8. Leyendo archivo Excel subido...")
        with pd.ExcelFile(budget_file) as xls:
            print(f"9. Hojas disponibles en el archivo: {xls.sheet_names}")
            
            if 'budget' not in xls.sheet_names:
                print("Error: No se encontró la hoja 'budget' en el archivo")
                return jsonify({"error": "El archivo no contiene la hoja 'budget'"}), 400
            
            print("10. Leyendo hoja 'budget'...")
            budget_data = pd.read_excel(xls, sheet_name='budget')
            print(f"11. Datos leídos. Dimensiones: {budget_data.shape}")
        
        # 5. Procesar el forecast_base.xlsm
        print("12. Cargando forecast_base.xlsm...")
        wb = load_workbook(forecast_path, read_only=False, keep_vba=True)
        print(f"13. Hojas disponibles en forecast: {wb.sheetnames}")
        
        # Verificar si existe la hoja oculta
        if 'Data Input Consol Q' not in wb.sheetnames:
            print("Error: No se encontró la hoja 'Data Input Consol Q'")
            return jsonify({"error": "No se encontró la hoja 'Data Input Consol Q'"}), 400
        
        # Acceder a la hoja (aunque esté oculta)
        print("14. Accediendo a la hoja 'Data Input Consol Q'...")
        ws = wb['Data Input Consol Q']
        print(f"15. Hoja obtenida. Filas actuales: {ws.max_row}, Columnas: {ws.max_column}")
        
        # 6. Limpiar hoja existente
        print("16. Limpiando hoja existente...")
        ws.delete_rows(1, ws.max_row)
        print(f"17. Hoja limpiada. Filas después de limpiar: {ws.max_row}")
        
        # 7. Escribir los nuevos datos
        print("18. Escribiendo nuevos datos...")
        for r in dataframe_to_rows(budget_data, index=False, header=True):
            ws.append(r)
        print(f"19. Datos escritos. Filas después de escribir: {ws.max_row}")
        
        # 8. Guardar cambios
        print("20. Guardando cambios en forecast_base.xlsm...")
        wb.save(forecast_path)
        print("21. Archivo guardado exitosamente")
        
        return jsonify({"message": "Archivo procesado correctamente"}), 200
    
    except Exception as e:
        print(f"ERROR CRÍTICO: {str(e)}")
        logging.error(f"Error al procesar archivo budget: {str(e)}", exc_info=True)
        return jsonify({
            "error": "Error al procesar el archivo",
            "details": str(e),
            "step": "Ver logs para ver el último paso ejecutado antes del error"
        }), 500


def wait_for_download(timeout=120):
    """Espera hasta que finalice la descarga con verificación mejorada"""
    start_time = time.time()
    last_check = start_time
    
    while (time.time() - start_time) < timeout:
        # Verificar archivos .crdownload cada 1 segundo
        if not any(f.endswith('.crdownload') for f in os.listdir(DOWNLOAD_FOLDER)):
            return True
        
        # Mostrar progreso cada 15 segundos
        if (time.time() - last_check) > 15:
            remaining = int(timeout - (time.time() - start_time))
            print(f"Descargando... Tiempo restante: {remaining}s")
            last_check = time.time()
            
        time.sleep(1)
    
    raise TimeoutError(f"La descarga no completó en {timeout} segundos")

@main.route('/run-process', methods=['POST'])
def run_process():
    driver = None
    try:
        # 1. Configurar navegador
        driver = webdriver.Edge(
            service=EdgeService(EDGE_DRIVER_PATH),
            options=configure_edge()
        )
        driver.implicitly_wait(10)

        # 2. Navegación y login (COMPLETA LOS XPATH)
        driver.get(COGNOS_URL)
        
        # Paso 1: Selección de opción
        WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="downshift-0-toggle-button"]'))
        ).click()
        
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="downshift-0-item-1"]'))
        ).click()

        WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="credsDiv"]'))
        ).click()

        # Paso 2: Login
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="user-name-input"]'))
        ).send_keys(COGNOS_USER)
        
        driver.find_element(By.XPATH, '//*[@id="password-input"]').send_keys(COGNOS_PASS)
        driver.find_element(By.XPATH, '//*[@id="login-button"]').click()

        
        WebDriverWait(driver, 50).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="com.ibm.bi.content.navigator.teamContent__panel"]/div[2]/div[2]/div/div[1]/div/label[3]/span[2]/div/div[1]/div/a/div/div'))
        ).click()
        time.sleep(10)  # Esperar manualmente si es necesario
        
         # Ahora puedes acceder al div con scroll
        scrollable_div = driver.find_element(By.XPATH, '//*[@id="com.ibm.bi.content.navigator.teamContent__panel"]/div[2]/div[2]/div')
        driver.execute_script("arguments[0].scrollTop = 800", scrollable_div)

        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="com.ibm.bi.content.navigator.teamContent__panel"]/div[2]/div[2]/div/div[1]/div/label[29]/span[2]/div/div[1]/div/a/div/div'))
        ).click()

        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="com.ibm.bi.content.navigator.teamContent__panel"]/div[2]/div[2]/div/div[1]/div/label[3]/span[2]/div/div[1]/div/a/div/div'))
        ).click()

        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="com.ibm.bi.content.navigator.teamContent__panel"]/div[2]/div[2]/div/div[1]/div/label[2]/span[2]/div/div[1]/div/a/div/div'))
        ).click()

        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="com.ibm.bi.content.navigator.teamContent__panel"]/div[2]/div[2]/div/div[1]/div/label[1]/span[2]/div/div[1]/div/a/div/div'))
        ).click()

        time.sleep(5)  # Esperar manualmente si es necesario
        
         # Ahora puedes acceder al div con scroll
        scrollable_div = driver.find_element(By.XPATH, '//*[@id="com.ibm.bi.content.navigator.teamContent__panel"]/div[2]/div[2]/div')
        driver.execute_script("arguments[0].scrollTop = 150", scrollable_div)

        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="com.ibm.bi.content.navigator.teamContent__panel"]/div[2]/div[2]/div/div[2]/div/label[4]/span[2]/div/div[1]/div/a/div/div/div/div'))
        ).click()
        time.sleep(50)  # Esperar manualmente si es necesario

        print("1. Esperando el iframe rsIFrameManager_1...")
        WebDriverWait(driver, 50).until(
            EC.frame_to_be_available_and_switch_to_it((By.XPATH, '//*[@id="rsIFrameManager_1"]'))
        )
        print("2. Cambio al iframe exitoso")

        print("3. Buscando el div scrollable clsViewerPage...")
        scrollable_div = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "clsViewerPage"))
        )
        print(f"4. Div encontrado - Visible: {scrollable_div.is_displayed()} - Tamaño: {scrollable_div.size}")

        print("5. Ejecutando scroll...")
        driver.execute_script("arguments[0].scrollTop = 400", scrollable_div)

        # Verificación
        current_scroll = driver.execute_script("return arguments[0].scrollTop", scrollable_div)
        print(f"6. Scroll completado - Posición actual: {current_scroll}")

        time.sleep(5) # Esperar a que se cargue el contenido

        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__17"]'))
        ).click()

        driver.switch_to.default_content()

        print("1. Esperando el iframe rsIFrameManager_1...")
        WebDriverWait(driver, 50).until(
            EC.frame_to_be_available_and_switch_to_it((By.XPATH, '//*[@id="rsIFrameManager_1"]'))
        )
        print("2. Cambio al iframe exitoso")

        print("3. Buscando el div scrollable clsViewerPage...")
        scrollable_div = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CLASS_NAME, "clsViewerPage"))
        )
        print(f"4. Div encontrado - Visible: {scrollable_div.is_displayed()} - Tamaño: {scrollable_div.size}")

        print("5. Ejecutando scroll...")
        driver.execute_script("arguments[0].scrollTop = 2100", scrollable_div)

        # Verificación
        current_scroll = driver.execute_script("return arguments[0].scrollTop", scrollable_div)
        print(f"6. Scroll completado - Posición actual: {current_scroll}")

        time.sleep(5) # Esperar a que se cargue el contenido

        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__163"]/td/div'))
        ).click()#Opp Age
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__164"]/td/div'))
        ).click()#Opportunity Source Name
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__165"]/td/div'))
        ).click()#Opportunity Source Code
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__166"]/td/div'))
        ).click()#Opp Owner Name
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__167"]/td/div'))
        ).click()#Opp Owner Email
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__168"]/td/div'))
        ).click()#Product Family Name
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__169"]/td/div'))
        ).click()#Product Family Code
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__170"]/td/div'))
        ).click()#Classification Name
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__172"]/td/div'))
        ).click()#Term
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__173"]/td/div'))
        ).click()#Renewal Type Name
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__174"]/td/div'))
        ).click()#DS Focus Products
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__175"]/td/div'))
        ).click()#Owning Org Channel Code
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__176"]/td/div'))
        ).click()#Owning Org Channel Name
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__177"]/td/div'))
        ).click()#Sales Motion Name
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__178"]/td/div'))
        ).click()#Channel Name
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__179"]/td/div'))
        ).click()#Business Partner Name
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__180"]/td/div'))
        ).click()#Business Partner Code
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__181"]/td/div'))
        ).click()#Offering GTM Status
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__182"]/td/div'))
        ).click()#Sales Competitors Name
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__183"]/td/div'))
        ).click()#Opp Identifier
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__184"]/td/div'))
        ).click()#Opp Creator Name
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__185"]/td/div'))
        ).click()#Opp Next Step
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__186"]/td/div'))
        ).click()#Type Code
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__187"]/td/div'))
        ).click()#Type Name
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__188"]/td/div'))
        ).click()#Deal Size
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__189"]/td/div'))
        ).click()#SS Last Modified Date
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__190"]/td/div'))
        ).click()#SS Elapsed Days
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__191"]/td/div'))
        ).click()#Mon1 Signings Rev
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__192"]/td/div'))
        ).click()#Mon2 Signings Rev
        WebDriverWait(driver, 90).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="hal__dom__uniqueID__193"]/td/div'))
        ).click()#Mon3 Signings Rev
  
        driver.switch_to.default_content()

        print("1. Esperando el iframe rsIFrameManager_1...")
        WebDriverWait(driver, 50).until(
            EC.frame_to_be_available_and_switch_to_it((By.XPATH, '//*[@id="rsIFrameManager_1"]'))
        )
        print("2. Cambio al iframe exitoso")

        print("3. Buscando el div scrollable clsViewerPage...")
        scrollable_div = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CLASS_NAME, "clsViewerPage"))
        )
        print(f"4. Div encontrado - Visible: {scrollable_div.is_displayed()} - Tamaño: {scrollable_div.size}")

        print("5. Ejecutando scroll...")
        driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", scrollable_div)

        # Verificación
        current_scroll = driver.execute_script("return arguments[0].scrollTop", scrollable_div)
        print(f"6. Scroll completado - Posición actual: {current_scroll}")

        time.sleep(20) # Esperar a que se cargue el contenido

        # Hacer clic en el botón Finalizar
        WebDriverWait(driver, 50).until(
            EC.element_to_be_clickable((By.XPATH, '/html/body/div[30]/div/div[1]/div[4]/div/table/tbody/tr[3]/td/button[2]'))
        ).click()


        time.sleep(100)  # Esperar a que se cargue el contenido

        WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="AppToolbarLeftPane_btnFormats"]'))
        ).click()#Clic a bton Html
        time.sleep(5)  # Esperar a que se cargue el contenido
        WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="mnuAppToolbarLeftPaneFormats_spreadsheetML"]/tbody/tr/td[2]'))
        ).click()#Clic a bton Excel

        time.sleep(100)  # Espera corta para que comience la descarga
    
        # Espera inicio de descarga con verificación progresiva
        download_started = False
        for attempt in range(1, 11):  # 10 intentos con espera adaptable
            if any(f.endswith('.crdownload') for f in os.listdir(DOWNLOAD_FOLDER)):
                download_started = True
                break
    
             # Espera progresiva: 1s, 2s, 3s,..., hasta 4s máximo por intento
            wait_time = min(attempt, 4)
            print(f"Intento {attempt}/10 - Esperando {wait_time}s...")
            time.sleep(wait_time)

        if not download_started:
            print("¡Atención! No se detectó inicio de descarga. Verifica:")
            print("- Conexión a internet")
            print("- Que el archivo exista")
            print("- Permisos de escritura en la carpeta")
        else:
            # Espera finalización con timeout generoso
            wait_for_download(180)  # 3 minutos para completar
            print("✓ Descarga finalizada correctamente")
    finally:
        if driver:
            driver.quit()

    def get_latest_file(folder):
        files = [os.path.join(folder, f) for f in os.listdir(folder) if f.endswith(('.xlsx', '.xlsm', '.xls'))]
        if not files:
            raise FileNotFoundError("No se encontró ningún archivo .xlsx")
        return max(files, key=os.path.getctime)
    try:
        archivo_descargado = get_latest_file(DOWNLOAD_FOLDER)

        # Validar que se descargó
        if not os.path.exists(archivo_descargado):
            raise FileNotFoundError("No se encontró el archivo descargado")
        # Ruta forecast base
        forecast_base = os.path.join(BASE_DIR, 'data', 'forecast_base.xlsx')
       
        # Procesar y enviar el resultado
        resultado = procesar_forecast(archivo_descargado, forecast_base)
        return send_file(
        resultado,
        as_attachment=True,
        download_name="Systems_HW_North_SSA_EPM_ISC.xlsm"
        )
    except Exception as e:
        logging.error(f"Error al procesar el archivo descargado: {str(e)}", exc_info=True)
        return jsonify({
            "error": "No se pudo procesar el archivo",
            "details": str(e)
        }), 500
